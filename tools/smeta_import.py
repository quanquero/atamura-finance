# -*- coding: utf-8 -*-
"""Импорт сметы объекта (лист ППЗ ПФ) → статьи × блок × бюджет × помесячный план.
Формат ATAMŪRA: ВДЦ/ППЗ/ГПР (пример — смета ЖК Аура Алатау, 3 очередь, блоки 5-8).
Лист «ППЗ ПФ»: № | Наименование | Ед | Итого дог.цена с НДС | <месяцы 2026…2027>.
Блок берётся из названия раздела/строки («… Блок №5»); статья наследует текущий блок.

    python3 tools/smeta_import.py <файл.xlsx> --object "Аура" --ochered 3 [--dry|--post]

--dry — печать структуры (по умолчанию); --post — отправить на /api/smeta (X-Service-Key).
Требует openpyxl."""
import sys, os, re, json

BLOCK_RE = re.compile(r"[Бб]лок\s*№?\s*(\d+)")
CODE_RE = re.compile(r"^\d+(\.\d+)*$")


def parse_ppz(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["ППЗ ПФ"] if "ППЗ ПФ" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    # строка с месяцами
    hdr = next((i for i, r in enumerate(rows)
                if r and any(str(x).strip().lower() == "январь" for x in r if x)), None)
    if hdr is None:
        return [], []
    year_row = rows[hdr - 1] if hdr else []
    months = []                                    # [(col, 'YYYY-MM')]
    cur_year = None
    for j, val in enumerate(rows[hdr]):
        if year_row and j < len(year_row) and year_row[j] and re.search(r"20\d\d", str(year_row[j])):
            cur_year = re.search(r"(20\d\d)", str(year_row[j])).group(1)
        nm = str(val).strip().lower() if val else ""
        MON = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06",
               "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
        if nm in MON and cur_year:
            months.append((j, "%s-%s" % (cur_year, MON[nm])))
    recs, block = [], ""
    for r in rows[hdr + 1:]:
        name = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        code = str(r[0]).strip() if r[0] is not None else ""
        mb = BLOCK_RE.search(name)
        if mb:
            block = mb.group(1)
        if not CODE_RE.match(code) or not name:
            continue
        total = r[3] if len(r) > 3 and isinstance(r[3], (int, float)) else 0.0
        plan = {mm: float(r[j]) for j, mm in months if j < len(r) and isinstance(r[j], (int, float)) and r[j]}
        if total or plan:
            art = BLOCK_RE.sub("", name).strip(" ,")
            recs.append({"block": block, "code": code, "article": art,
                         "budget": float(total or 0), "plan": plan})
    return recs, [m[1] for m in months]


def main():
    args = sys.argv[1:]
    files = [a for a in args if a.lower().endswith(".xlsx")]
    if not files:
        print("Использование: python3 tools/smeta_import.py <файл.xlsx> --object Аура --ochered 3 [--post]")
        return
    obj = _opt(args, "--object", "")
    och = _opt(args, "--ochered", "")
    recs, months = parse_ppz(files[0])
    # только строки-статьи (с точкой в коде) — разделы без суммы отсекаем как заголовки
    arts = [r for r in recs if "." in r["code"]]
    print("Объект: %s · очередь: %s · месяцев плана: %d · статей: %d" % (obj or "?", och or "?", len(months), len(arts)))
    by_block = {}
    for r in arts:
        by_block.setdefault(r["block"] or "?", []).append(r)
    for blk, lst in sorted(by_block.items()):
        s = sum(x["budget"] for x in lst)
        print("\n  Блок №%s — бюджет %s ₸" % (blk, _m(s)))
        for x in sorted(lst, key=lambda z: -z["budget"])[:12]:
            print("    %-32s %14s" % (x["article"][:32], _m(x["budget"])))
    if "--post" in args:
        payload = {"object": obj, "ochered": och, "months": months, "articles": arts}
        _post(payload)


def _opt(args, key, default):
    return args[args.index(key) + 1] if key in args and args.index(key) + 1 < len(args) else default


def _m(n):
    return ("{:,.0f}".format(n or 0)).replace(",", " ")


def _post(payload):
    import urllib.request
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    import server as S
    if not S.KEY:
        print("SERVICE_KEY не задан"); return
    url = "https://finance.atamura.group/api/smeta"
    req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"),
                                 headers={"Content-Type": "application/json", "X-Service-Key": S.KEY})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            print("Отправлено:", json.load(r))
    except Exception as e:
        print("Ошибка отправки:", e)


if __name__ == "__main__":
    main()
