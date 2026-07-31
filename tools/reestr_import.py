# -*- coding: utf-8 -*-
"""
Импорт реестра финотдела из Excel → финсервер (POST /api/reestr).

Использование:
    SERVICE_KEY=... python tools/reestr_import.py "путь/к/реестру.xlsx"
    # по умолчанию шлёт на https://finance.atamura.group; переопределить: FIN_URL=...

Парсит Лист1 (построчный реестр: № заявки + поставщик/БИН/ИБАН + назначение)
и Лист5 (запросы наличных). Разбирает ячейку поставщика на имя/БИН/ИБАН.
"""
import sys, os, re, json, urllib.request
from datetime import datetime
import openpyxl

BIN = re.compile(r"\b(\d{12})\b")
IBAN = re.compile(r"\b(KZ[0-9A-Z]{18})\b")
NUM = re.compile(r"(\d{4,6})")


def _num(v):
    try:
        f = float(v)
        return f if f > 0 else 0.0
    except Exception:
        return 0.0


def _supplier(cell):
    s = str(cell or "").strip()
    b, ib = BIN.search(s), IBAN.search(s)
    name = s[:b.start()].rstrip(" ,").strip() if b else s
    return name, (b.group(1) if b else ""), (ib.group(1) if ib else "")


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = []
    if "Лист1" in wb.sheetnames:
        for r in wb["Лист1"].iter_rows(values_only=True):
            r = [("" if v is None else v) for v in r] + [""] * 10
            znum, (name, b, ib) = str(r[7]).strip(), _supplier(r[2])
            if not (NUM.fullmatch(znum)) and not b:
                continue
            rows.append({"src": "Лист1", "num": znum if NUM.fullmatch(znum) else "",
                         "name": name, "bin": b, "iban": ib, "amount": _num(r[4]) or _num(r[6]),
                         "purpose": str(r[8]).strip(), "invoice": str(r[9]).strip()})
    if "Лист5" in wb.sheetnames:
        for r in wb["Лист5"].iter_rows(values_only=True):
            r = [("" if v is None else v) for v in r] + [""] * 8
            znum = str(r[5]).strip()
            if not NUM.fullmatch(znum):
                continue
            rows.append({"src": "Лист5(нал)", "num": znum, "name": str(r[4]).strip(),
                         "bin": "", "iban": "", "amount": _num(r[3]),
                         "purpose": str(r[6]).strip(), "invoice": ""})
    return rows


def main():
    if len(sys.argv) < 2:
        print("Укажи путь к xlsx реестра"); sys.exit(1)
    rows = parse(sys.argv[1])
    nums = sorted({r["num"] for r in rows if r["num"]})
    bins = sorted({r["bin"] for r in rows if r["bin"]})
    print(f"Разобрано строк: {len(rows)} · № заявок: {len(nums)} · БИН: {len(bins)}")
    url = os.environ.get("FIN_URL", "https://finance.atamura.group").rstrip("/") + "/api/reestr"
    key = os.environ.get("SERVICE_KEY", "")
    if not key:
        print("Нет SERVICE_KEY — только парсинг, без отправки."); return
    payload = json.dumps({"rows": rows, "ts": datetime.now().strftime("%Y-%m-%d %H:%M")}).encode("utf-8")
    req = urllib.request.Request(url, data=payload,
                                 headers={"X-Service-Key": key, "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        print("Ответ сервера:", r.read().decode("utf-8"))


if __name__ == "__main__":
    main()
